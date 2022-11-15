# -*- coding: utf-8 -*-
"""
Created on Sun Mar 15 15:06:34 2020

@author: janak
"""

import os
import glob
import pandas as pd
import numpy as np
from scipy.spatial import distance

#Input pixel size as calculated from scale
#pixel_size = float(input('Define pixel size (nm/pixel)'))
pixel_size = 0.7059

#Creation of lists for all parameters that will be appended with every txt file 
#and saved as excel files in the end
az_length = []
N_docked = []
diameter_all = []
D_avg = []
N_ELV = []
A_ELV = []
SVsurface = []
AZ_counter = []


os.chdir('D:/Jana/Studenten/EM-Kurs/analysis/Lia')
print(os.getcwd())
filenames = (glob.glob('*.txt'))

for f in filenames:
    #Import of txt files generated with ImageJ, addition of headings in data frame
    test = pd.read_csv(f, delimiter='\t', names = ["#", "category", "value1", "value2", "value3"])
    test.set_index("category", inplace=True)
    
    AZ_number = (test.index == 'Active_zone').sum()
    
    if AZ_number == 1: 
        AZ_counter.append(1)
    else:
        print ('Problem with active zone in file:', f)

if(len(filenames)) == len(AZ_counter):
    print ('all fine')

    pi = 3.14159265359
    
    #bins for the binning of distances
    bins = np.arange(0,1001,10,int)  
    binningvalues = bins.tolist() 
    
    
    #creation of xls-File for distances
    from openpyxl.workbook import Workbook
    from openpyxl import load_workbook
    
    #headers = [filenames]
    workbook_name1 = 'distances.xlsx'
    wb = Workbook()
    page = wb.active
    page.title = 'distances'
    page.append(['hello','hello'])
    wb.save(filename = workbook_name1)
    
    #headers = [filenames]
    workbook_name2 = 'bins.xlsx'
    wb = Workbook()
    page = wb.active
    page.title = 'bins'
    page.append(binningvalues)
    wb.save(filename = workbook_name2)
    
    #headers = [filenames]
    workbook_name3 = 'neighbor.xlsx'
    wb = Workbook()
    page = wb.active
    page.title = 'neighbor'
    page.append(['hello','hello'])
    wb.save(filename = workbook_name3)
    
    #page.append(headers) # write the headers to the first line
    
    
    for f in filenames:
        #Import of txt files generated with ImageJ, addition of headings in data frame
        test = pd.read_csv(f, delimiter='\t', names = ["#", "category", "value1", "value2", "value3"])
        test.set_index("category", inplace=True)
        
        """
        Part 1 - Active zone
        """
        
        #Addition of active zone length to the list az_length
        length = (test.loc['Active_zone','value1'])
        az_length.append(length*pixel_size)
        
        #Creation of array with coordinates of all active zone values (x and y)
        x_coord = (test.loc['Active_zone','value2'])
        x = (x_coord.split(','))
        y_coord = (test.loc['Active_zone','value3'])
        y = (y_coord.split(','))
        xy_coord = np.array((x,y),float)
        xy_coord = np.transpose(xy_coord)
    
        """
        Part 2 - Vesicles (SV)
        """
        #Creation of data frame with only SV values
        sv_values = (test.loc['SV'])
        
        #Addition of all radiuses to list radius_all_sv (internal calculation),
        #to list radius_all (to collect all radiuses from all txt files)
        radius_all_sv = np.array(sv_values['value3'], float)
        radius_all_sv = radius_all_sv*pixel_size
        
        #Calculation surface area for all SVs in one image
        surfaces = 4 * pi * radius_all_sv ** 2
        SVsurface.append(surfaces.sum())
        
        #Creation of array with coordinates of all SVs 
        x_sv = (sv_values['value1'].astype(float))
        y_sv = (sv_values['value2'].astype(float))
        xy_sv = np.array((x_sv,y_sv))
        xy_sv = np.transpose(xy_sv)
        #print(xy_sv)
        
        #Calculation of min. distance between each SV and active zone
        #thereby subtraction of radius of each SV to have the distance between membranes
        all_dist = (distance.cdist(xy_sv,xy_coord).min(axis=1))*pixel_size-radius_all_sv
        #print(all_dist)
        
        #binning of distances
        out = pd.cut(all_dist, bins=bins)
        distancebins = pd.value_counts(out, sort = False)
        #.sort_index()
    
        #Calculation of nearest neighbor SV distance for nearest three SVs, 
        #thereby exclusion of docked SVs
        neighbor = (distance.cdist(xy_sv,xy_sv))
        neighborsorted = np.sort(neighbor)
        averageneighbor = (neighborsorted[:,1:4].mean(axis=1)*pixel_size)
        #print(averageneighbor)
        
        """
        Part 3 - Docked vesicles
        """
        #test how many docked SVs are listed in txt file
        m = (test.index == 'docked_SV').sum()
        
        if m == 0:
            N_docked.append(0)
            D = radius_all_sv*2
            diameter_all.append(D)
            D_avg.append(D.mean())
            
        elif m == 1:
            N_docked.append(1)
            docked_values = (test.loc['docked_SV'])
            radius_all_docked = np.array(docked_values['value3'],float)
            radius_all_docked = radius_all_docked*pixel_size
            D = np.append(radius_all_sv,radius_all_docked)*2
            diameter_all.append(D)
            D_avg.append(D.mean())
            
        else:
            #Creation of data frame with only docked_SV values
            docked_values = (test.loc['docked_SV'])
            
            #Addition of all radiuses to list radius_all_docked
            radius_all_docked = np.array(docked_values['value3'], float)
            radius_all_docked = radius_all_docked*pixel_size
            
            #Count of docked SVs
            N_docked.append(len(radius_all_docked))
            
            #Addition of all SV diameters to list diameter_all, addition of avg. D to D_avg
            D = np.concatenate([radius_all_sv,radius_all_docked])*2
            diameter_all.append(D)
            D_avg.append(D.mean())
            
        """
        Part 4 - Endosome-like vacuoles (ELVs)
        
        -- still missing: distance of ELVs to AZ --
        """
        #test how many endosomes are listed in txt file
        n = (test.index == 'Endosome').sum()
        
        if n == 0:
            N_ELV.append(0)
            A_ELV.append(0)
            
        elif n == 1:
            N_ELV.append(1)
            ELV_values = (test.loc['Endosome'])
            ELV_area = (ELV_values['value1'])
            ELV_area = ELV_area*pixel_size*pixel_size
            A_ELV.append(ELV_area)
        
        else:
            
            #Creation of data frame with only endosome values
            ELV_values = (test.loc['Endosome'])
            
            #Creation of array with ELV areas
            ELV_area = np.array(ELV_values['value1'], float)
            ELV_area = ELV_area*pixel_size*pixel_size
            
            #Addition of ELV number to list N_ELV
            N_ELV.append(len(ELV_area))
            A_ELV.append(ELV_area.mean())
            
        # """
        # Part 5 - Export as Excel files
        # """
        
        wb = load_workbook(workbook_name1)
        page = wb.active
        dist = all_dist.tolist()
        page.append(dist)
        wb.save(filename = workbook_name1)
        wb.close()
        
        wb = load_workbook(workbook_name2)
        page = wb.active
        binning = distancebins.tolist()
        page.append(binning)
        wb.save(filename = workbook_name2)
        wb.close()
      
        wb = load_workbook(workbook_name3)
        page = wb.active
        neighbors = averageneighbor.tolist()
        page.append(neighbors)
        wb.save(filename = workbook_name3)
        wb.close()   
            
    print (len(filenames))
    print (filenames)
    
    alltogether = pd.DataFrame(list(zip(filenames, az_length, N_docked, D_avg, N_ELV, A_ELV,SVsurface)), columns =['File', 'AZ_length', 'N_docked', 'AVG_SV_diameter', 'N_ELVs', 'Avg_ELV_area', 'Summed SV surface area']) 
    print (alltogether)
    
    alltogether.to_excel(r'D:/Jana/Studenten/EM-Kurs/analysis/Lia/summary_surface.xlsx', index = False)
    
else:
    print('Oopsie')
       